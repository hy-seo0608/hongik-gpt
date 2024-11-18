import openpyxl
import os

def replace_marker_and_save_new_file(excel_file_path, renew_excel_file_path, new_file_path):
    # 엑셀 파일 열기
    workbook = openpyxl.load_workbook(excel_file_path)
    stdbook = openpyxl.load_workbook(renew_excel_file_path)
    sheet = workbook.active  # 활성 시트를 선택합니다. (첫 번째 시트)
    std_sheet = stdbook.active
    # 모든 셀을 순회하면서 $$ 표식이 있는지 확인
    for row in std_sheet.iter_rows() : 
        marker = row[0].value
        replacement_text = str(row[1].value)
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and marker in cell.value:
                    # 표식을 찾아서 텍스트로 대체
                    cell.value = cell.value.replace(marker, replacement_text)

    # 수정된 내용을 새 파일로 저장
    workbook.save(new_file_path)

print(os.getcwd())
std_excel_file_path = '../../dataset/answerfile_template.xlsx'         # 원본 엑셀 파일 경로
renew_excel_file_path = '../../dataset/renew.xlsx'         # 크롤링 후 갱신하는 파일 경로
new_file_path = '../../dataset/answerfile.xlsx'       # 새롭게 저장할 엑셀 파일 경로

replace_marker_and_save_new_file(std_excel_file_path, renew_excel_file_path, new_file_path)