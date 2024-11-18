# from utils.scrapper import Scrapper
import openpyxl

# renew.xlsx 파일 열기
def fill_excel_with_scraped_data(excel_file_path):
    # 엑셀 파일 불러오기
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active  # 활성화된 시트 선택

    # 스크래핑된 데이터 가져오기
    

    # 엑셀 파일에서 'text'라는 열을 찾아서 데이터 입력
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):  # 첫 번째 행은 헤더
        cell = row[sheet["A1"].column - 1]  # 'text' 열이 A열이라고 가정
        if isinstance(cell.value, str) and "text" in cell.value.lower():
            # 셀에 스크래핑된 데이터를 삽입 (순서대로 삽입한다고 가정)
            cell.value = "여기다가 넣기"

    # 엑셀 파일 저장 (원본 파일에 덮어쓰기)
    workbook.save(excel_file_path)

renew_file_path = '/mysite/dataset/renew.xlsx'
fill_excel_with_scraped_data(renew_file_path)

import datetime

# 오늘 날짜 계산
today = datetime.date.today()

today_year = today.year
today_month = today.month
today_day = today.day
today_week = today.strftime("%A")  # 요일은 문자열로 반환 ('Monday', 'Tuesday' 등)

# 내일 날짜 계산
tomorrow = today + datetime.timedelta(days=1)

tomorrow_year = tomorrow.year
tomorrow_month = tomorrow.month
tomorrow_day = tomorrow.day
tomorrow_week = tomorrow.strftime("%A")  # 내일의 요일

# 결과 출력 (확인용)
print(f"오늘: {today_year}-{today_month}-{today_day} ({today_week})")
print(f"내일: {tomorrow_year}-{tomorrow_month}-{tomorrow_day} ({tomorrow_week})")