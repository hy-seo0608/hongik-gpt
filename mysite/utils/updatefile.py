import schedule
import datetime
import pandas as pd
from .scrapper import Scrapper
import json
from collections import OrderedDict
import requests
import openpyxl
import os
from .weather import get_weather
from .api import fetch_and_save_json

import sys

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
from configure import RENEW_FILE_PATH, FOOD_LIST_FILE_PATH, API_FOOD_URL, NOTICE_FILE_PATH

# Scrapper 클래스 인스턴스 생성
scraper = Scrapper()

# 요일을 한국어로 매핑하는 딕셔너리
weekdays_korean = {"Monday": "월", "Tuesday": "화", "Wednesday": "수", "Thursday": "목", "Friday": "금", "Saturday": "토", "Sunday": "일"}


# 오늘, 내일 날짜 업데이트
def update_date(excel_file_path):
    today = datetime.date.today()

    today_year = today.year
    today_month = today.month
    today_day = today.day
    today_week = weekdays_korean[today.strftime("%A")]  # 요일을 한국어로 변환

    # 내일 날짜 계산
    tomorrow = today + datetime.timedelta(days=1)

    tomorrow_year = tomorrow.year
    tomorrow_month = tomorrow.month
    tomorrow_day = tomorrow.day
    tomorrow_week = weekdays_korean[tomorrow.strftime("%A")]  # 내일의 요일을 한국어로 변환
    # 엑셀 파일 불러오기
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active  # 활성화된 시트 선택

    sheet["B8"] = today_year
    sheet["B9"] = today_month
    sheet["B10"] = today_day
    sheet["B11"] = today_week
    sheet["B19"] = tomorrow_year
    sheet["B20"] = tomorrow_month
    sheet["B21"] = tomorrow_day
    sheet["B22"] = tomorrow_week

    workbook.save(RENEW_FILE_PATH)


# 학식 정보 크롤링 및 엑셀 파일 업데이트
def update_food_list(excel_file_path):
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active  # 활성화된 시트 선택

    fetch_and_save_json(API_FOOD_URL, FOOD_LIST_FILE_PATH)

    # JSON 파일 로드
    with open(FOOD_LIST_FILE_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)
    # 각 섹션을 개별 데이터프레임으로 변환
    df = pd.DataFrame(data["DATA"]["RESTDATA"])

    def find_menu(menu_date, rest_no, price_level):
        # 조건 필터링
        filtered = df[(df["MENU_DATE"] == menu_date) & (df["REST_NO"] == rest_no) & (df["PRICELEVEL"] == price_level)]
        if not filtered.empty:
            lst = list(filtered.iloc[0]["MENU"].split("\r\n"))
            lst = list(filter(bool, lst))
            return ", ".join(lst)  # 첫 번째 결과 반환
        else:
            return "메뉴를 찾을 수 없습니다. "

    today_date = datetime.datetime.now().strftime("%Y%m%d")
    tomorrow_date = (datetime.datetime.now() + datetime.timedelta(days=1)).strftime("%Y%m%d")
    staff_no, domi_no = "2", "3"
    # 리스트를 문자열로 변환하여 엑셀에 저장
    # 아침 : 0 점심 : 1 점심A : 2 저녁 : 3
    sheet["B2"] = find_menu(today_date, domi_no, "0")
    sheet["B3"] = find_menu(today_date, domi_no, "1")
    sheet["B28"] = find_menu(today_date, domi_no, "2")
    sheet["B4"] = find_menu(today_date, domi_no, "3")

    sheet["B5"] = find_menu(today_date, staff_no, "3")
    sheet["B6"] = find_menu(today_date, staff_no, "0")
    sheet["B7"] = find_menu(today_date, staff_no, "1")

    # 다음날
    sheet["B29"] = find_menu(tomorrow_date, domi_no, "0")
    sheet["B30"] = find_menu(tomorrow_date, domi_no, "1")
    sheet["B31"] = find_menu(tomorrow_date, domi_no, "2")
    sheet["B32"] = find_menu(tomorrow_date, domi_no, "3")

    sheet["B33"] = find_menu(tomorrow_date, staff_no, "3")
    sheet["B34"] = find_menu(tomorrow_date, staff_no, "0")
    sheet["B35"] = find_menu(tomorrow_date, staff_no, "1")

    workbook.save(RENEW_FILE_PATH)


# 공지사항 크롤링 및 엑셀 파일 업데이트
def update_notice(excel_file_path):
    scraper.get_notice()

    # 공지사항 중 중복되는 것 없애기
    with open(NOTICE_FILE_PATH, "r") as f:
        notices = json.load(f)
    # 중복되는 공지 링크 기준으로 제거
    seen_links = set()
    unique_notices = []
    for notice in notices:
        if notice["link"] not in seen_links:
            unique_notices.append(notice)
            seen_links.add(notice["link"])
    # 중복 제거된 공지사항들 json 파일에 기록
    with open(NOTICE_FILE_PATH, "w") as f:
        json.dump(unique_notices, f, ensure_ascii=False, indent="\t")

    notice_df = pd.DataFrame(unique_notices)
    ret_notice = ""
    for i in range(min(5, len(notice_df))):
        ret_notice += f"작성자 : {notice_df.iloc[i]['name']}\n"
        ret_notice += f"제목 : {notice_df.iloc[i]['subject']}\n"
        ret_notice += f"링크 : {notice_df.iloc[i]['link']}\n"
        ret_notice += f"작성일자 : {notice_df.iloc[i]['date']}\n\n"
    # 엑셀 파일 불러오기
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active  # 활성화된 시트 선택

    sheet["B12"] = ret_notice
    workbook.save(RENEW_FILE_PATH)


def update_weather(excel_file_path):
    today = datetime.datetime.now()
    today_weather = get_weather(today)
    tomorrow = today + datetime.timedelta(days=1)
    tomorrow_weather = get_weather(tomorrow)

    # 엑셀 파일 불러오기
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active  # 활성화된 시트 선택

    if today_weather != None:
        sheet["B14"] = today_weather["온도"]
        sheet["B15"] = today_weather["날씨"]
        sheet["B16"] = today_weather["최저온도"]
        sheet["B17"] = today_weather["최고온도"]
    else:
        print("Error fetching weather data")

    if tomorrow_weather != None:
        sheet["B23"] = tomorrow_weather["온도"]
        sheet["B24"] = tomorrow_weather["날씨"]
        sheet["B25"] = tomorrow_weather["최저온도"]
        sheet["B26"] = tomorrow_weather["최고온도"]
        workbook.save(RENEW_FILE_PATH)
    else:
        print("Error fetching weather data")


"""
# 날짜 정보는 하루에 한 번 업데이트
schedule.every().day.at("00:00").do(update_date)

# 학식 정보는 하루에 한 번 업데이트
schedule.every().day.at("07:00").do(update_food_list)

# 공지사항과 열람실 정보는 매 1시간마다 업데이트
schedule.every(1).hours.do(update_notice)
schedule.every(1).hours.do(uqdate_studyroom_status)                            

# 스크립트 무한 루프 실행
while True:
    schedule.run_pending()
    time.sleep(1)
"""

import os

if __name__ == "__main__":
    print(os.getcwd())
    update_date(RENEW_FILE_PATH)
    update_food_list(RENEW_FILE_PATH)
    update_notice(RENEW_FILE_PATH)
    update_weather(RENEW_FILE_PATH)

    # # 테스트용: 1분으로 설정하고 테스트
    # schedule.every().minutes.do(lambda: update_date(renewfile_path))
    # schedule.every().minutes.do(lambda : update_food_list(renewfile_path))
    # schedule.every(1).minutes.do(lambda : update_notice(renewfile_path))
    # schedule.every(1).minutes.do(lambda : update_weather(renewfile_path))
    # # schedule.every(1).minutes.do(update_studyroom_status)

    # # 테스트용: 모든 작업을 즉시 실행
    # schedule.run_all()
